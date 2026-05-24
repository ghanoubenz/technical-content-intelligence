from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

blue_fill = PatternFill('solid', fgColor='13499C')
grey_fill = PatternFill('solid', fgColor='F2F2F2')
header_font = Font(name='Montserrat', bold=True, color='FFFFFF', size=11)
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, cols):
    for col_num, header in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, num_cols, alt=False):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        if alt:
            cell.fill = grey_fill

# === SHEET 1: Content Calendar ===
ws1 = wb.active
ws1.title = 'Content Calendar'
cal_headers = ['Post ID', 'Week', 'Planned Date', 'Pillar', 'Mix Category', 'Topic',
               'Tool Link', 'Format', 'Hook', 'Status', 'Caption', 'Slides Link',
               'Approved By', 'Post URL', 'Email Reuse', 'Website Reuse', 'PDF Download']
style_header(ws1, cal_headers)

widths = [16, 8, 14, 14, 16, 35, 14, 14, 45, 12, 50, 18, 14, 25, 12, 14, 14]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

data = [
    ['POST-W01-01', 'W01', '2026-04-14', 'Educational', 'Problem-based', 'How MFL Actually Works',
     'MFL', 'Carousel', 'MFL is the most used ILI tech. But most people cant explain how it works.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W01-02', 'W01', '2026-04-16', 'Industry', 'Insight-based',
     'AI in Pipeline Inspection - Hype vs Reality', 'General', 'Single Post',
     'AI wont replace pipeline inspectors. But it is changing how we read the data.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W02-01', 'W02', '2026-04-21', 'Problem', 'Problem-based',
     'What Grows Between Inspections', 'UTWM + UTCD', 'Carousel',
     'Your last ILI was 3 years ago. The pipeline isnt the same pipe anymore.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W02-02', 'W02', '2026-04-23', 'Product', 'Tool-based',
     'UTCD vs MFL - When to Use Which', 'UTCD + MFL', 'Carousel',
     'MFL or UTCD? Wrong question.', 'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W02-03', 'W02', '2026-04-25', 'Event', 'Insight-based',
     'Team Spotlight - Field Engineers', 'General', 'Single Post',
     'Behind every ILI run is a team that prepared for months.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W03-01', 'W03', '2026-04-28', 'Educational', 'Problem-based',
     'What EMAT Detects and Why Its Different', 'EMAT', 'Carousel',
     'No liquid. No couplant. EMAT inspects gas pipelines differently.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W03-02', 'W03', '2026-04-30', 'Problem', 'Problem-based',
     'The 40% Unpiggable Problem', 'General', 'Single Post',
     '40% of the worlds pipelines cant run a conventional ILI tool.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W04-01', 'W04', '2026-05-05', 'Educational', 'Problem-based',
     'Why Cleaning Before ILI Changes Everything', 'Cleaning', 'Carousel',
     'Garbage in, garbage out. ILI accuracy starts before the tool enters the pipe.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W04-02', 'W04', '2026-05-07', 'Educational', 'Tool-based',
     'Pipeline Cracks Explained: SCC vs Fatigue', 'UTCD', 'Carousel',
     'Not all cracks behave the same. SCC, fatigue, seam weld - different behavior.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
    ['POST-W04-03', 'W04', '2026-05-09', 'Industry', 'Insight-based',
     'Hydrogen Pipeline Inspection Challenges', 'EMAT', 'Single Post',
     'Hydrogen is coming to pipelines. The inspection playbook hasnt caught up.',
     'IDEA', '', '', '', '', 'N', 'N', 'N'],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=val)
    style_row(ws1, row_idx, len(cal_headers), alt=(row_idx % 2 == 0))

status_dv = DataValidation(type='list', formula1='"IDEA,DRAFTED,DESIGNED,APPROVED,POSTED"', allow_blank=True)
ws1.add_data_validation(status_dv)
for r in range(2, 102):
    status_dv.add(ws1.cell(row=r, column=10))

pillar_dv = DataValidation(type='list', formula1='"Educational,Problem,Product,Industry,Event"', allow_blank=True)
ws1.add_data_validation(pillar_dv)
for r in range(2, 102):
    pillar_dv.add(ws1.cell(row=r, column=4))

format_dv = DataValidation(type='list', formula1='"Carousel,Single Post,Video Idea"', allow_blank=True)
ws1.add_data_validation(format_dv)
for r in range(2, 102):
    format_dv.add(ws1.cell(row=r, column=8))

mix_dv = DataValidation(type='list', formula1='"Problem-based,Tool-based,Insight-based"', allow_blank=True)
ws1.add_data_validation(mix_dv)
for r in range(2, 102):
    mix_dv.add(ws1.cell(row=r, column=5))

yn_dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
ws1.add_data_validation(yn_dv)
for r in range(2, 102):
    for c in [15, 16, 17]:
        yn_dv.add(ws1.cell(row=r, column=c))

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:Q{len(data)+1}'

# === SHEET 2: Performance ===
ws2 = wb.create_sheet('Performance')
perf_headers = ['Post ID', 'Date Posted', 'Impressions', 'Reactions', 'Comments',
                'Shares', 'Engagement Rate', 'Website Clicks', 'Save Rate',
                'Top Comment Theme', 'Learning']
style_header(ws2, perf_headers)
perf_widths = [16, 14, 14, 12, 12, 10, 16, 16, 12, 30, 40]
for i, w in enumerate(perf_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for row_idx in range(2, 12):
    post_id = data[row_idx - 2][0] if row_idx - 2 < len(data) else ''
    ws2.cell(row=row_idx, column=1, value=post_id)
    eng_formula = f'=IF(C{row_idx}>0,(D{row_idx}+E{row_idx}+F{row_idx})/C{row_idx},0)'
    ws2.cell(row=row_idx, column=7, value=eng_formula)
    ws2.cell(row=row_idx, column=7).number_format = '0.00%'
    style_row(ws2, row_idx, len(perf_headers), alt=(row_idx % 2 == 0))

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = 'A1:K11'

# === SHEET 3: Outreach Tracker ===
ws3 = wb.create_sheet('Outreach Tracker')
out_headers = ['Post ID', 'Content Piece', 'Used in Email', 'Campaign Name',
               'Email Opens', 'Replies', 'Meeting Booked', 'Notes']
style_header(ws3, out_headers)
out_widths = [16, 35, 14, 25, 14, 10, 16, 40]
for i, w in enumerate(out_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

yn_dv2 = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
ws3.add_data_validation(yn_dv2)
for r in range(2, 102):
    yn_dv2.add(ws3.cell(row=r, column=3))
    yn_dv2.add(ws3.cell(row=r, column=7))

ws3.freeze_panes = 'A2'

# === SHEET 4: Special Days ===
ws4 = wb.create_sheet('Special Days')
sd_headers = ['Date', 'Day Name', 'Region', 'Relevant?', 'Connection to Pipeline/Energy', 'Post Status']
style_header(ws4, sd_headers)
sd_widths = [14, 30, 14, 12, 45, 14]
for i, w in enumerate(sd_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

special_days = [
    ['2026-02-16', 'Engineers Week', 'Global', 'Y', 'Spotlight on pipeline engineering team', 'IDEA'],
    ['2026-03-04', 'World Engineering Day', 'Global', 'Y', 'Engineers behind ILI tools', 'IDEA'],
    ['2026-03-22', 'World Water Day', 'Global', 'Y', 'Pipeline integrity prevents water contamination', 'IDEA'],
    ['2026-04-28', 'World Day for Safety at Work', 'Global', 'Y', 'Pipeline safety and inspection culture', 'IDEA'],
    ['2026-06-05', 'World Environment Day', 'Global', 'Y', 'How inspection prevents pipeline leaks', 'IDEA'],
    ['2026-09-23', 'Saudi National Day', 'KSA', 'Y', 'Regional energy infrastructure support', 'IDEA'],
    ['2026-10-22', 'World Energy Day', 'Global', 'Y', 'Pipelines carry majority of world energy', 'IDEA'],
    ['2026-11-12', 'World Quality Day', 'Global', 'Y', 'Data quality in pipeline inspection', 'IDEA'],
    ['2026-12-02', 'UAE National Day', 'UAE', 'Y', 'Supporting UAE energy infrastructure', 'IDEA'],
]

for row_idx, row_data in enumerate(special_days, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=val)
    style_row(ws4, row_idx, len(sd_headers), alt=(row_idx % 2 == 0))

sd_status_dv = DataValidation(type='list', formula1='"SKIP,IDEA,DRAFTED,POSTED"', allow_blank=True)
ws4.add_data_validation(sd_status_dv)
for r in range(2, 52):
    sd_status_dv.add(ws4.cell(row=r, column=6))

ws4.freeze_panes = 'A2'

# === SHEET 5: Topic Bank ===
ws5 = wb.create_sheet('Topic Bank')
tb_headers = ['Topic', 'Pillar', 'Tool Link', 'Suggested Format', 'Hook Idea', 'Priority', 'Used?']
style_header(ws5, tb_headers)
tb_widths = [40, 14, 14, 16, 50, 12, 10]
for i, w in enumerate(tb_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

topics = [
    ['How MFL Actually Works', 'Educational', 'MFL', 'Carousel',
     'Most people cant explain what happens inside the pipe', 'HIGH', 'Y'],
    ['What Grows Between Inspections', 'Problem', 'UTWM+UTCD', 'Carousel',
     'Your last ILI was 3 years ago', 'HIGH', 'Y'],
    ['UTCD vs MFL Comparison', 'Product', 'UTCD+MFL', 'Carousel',
     'MFL or UTCD? Wrong question.', 'HIGH', 'Y'],
    ['What EMAT Detects', 'Educational', 'EMAT', 'Carousel',
     'No couplant. No liquid. Different physics.', 'HIGH', 'Y'],
    ['Why Cleaning Before ILI Matters', 'Educational', 'Cleaning', 'Carousel',
     'Garbage in garbage out', 'HIGH', 'N'],
    ['Types of Pipeline Cracks', 'Educational', 'UTCD', 'Carousel',
     'Not all cracks behave the same', 'HIGH', 'N'],
    ['The 40% Unpiggable Problem', 'Problem', 'General', 'Single Post',
     '40% of world pipelines cant run ILI', 'MEDIUM', 'N'],
    ['AI in Pipeline Inspection', 'Industry', 'General', 'Single Post',
     'AI wont replace inspectors', 'MEDIUM', 'N'],
    ['Hydrogen Pipeline Challenges', 'Industry', 'EMAT', 'Single Post',
     'Hydrogen is coming. Inspection isnt ready.', 'MEDIUM', 'N'],
    ['Corrosion: Internal vs External', 'Educational', 'MFL+UTWM', 'Carousel',
     'Same word. Different threat. Different tool.', 'MEDIUM', 'N'],
    ['Dent + Metal Loss Interaction', 'Problem', 'MFL', 'Single Post',
     'A dent alone is acceptable. Add corrosion and it changes.', 'MEDIUM', 'N'],
    ['What Happens During an ILI Run', 'Educational', 'General', 'Carousel',
     'Most people never see inside a pipeline during inspection', 'MEDIUM', 'N'],
    ['Predictive vs Reactive Integrity', 'Industry', 'General', 'Single Post',
     'Most operators are still reactive. Costs more.', 'LOW', 'N'],
    ['Digital Twins in Pipeline Integrity', 'Industry', 'General', 'Single Post',
     'Digital twins - practical or buzzword?', 'LOW', 'N'],
    ['UTWM: Precision Wall Measurement', 'Product', 'UTWM', 'Carousel',
     'Sub-millimeter accuracy. Why that matters.', 'MEDIUM', 'N'],
    ['Tool Selection Guide', 'Product', 'All', 'Carousel',
     'Picking the right ILI tool starts with one question', 'HIGH', 'N'],
    ['Pipeline Failures That Were Preventable', 'Problem', 'General', 'Single Post',
     'Every major pipeline failure has a lesson', 'MEDIUM', 'N'],
    ['Robotic ILI Solutions', 'Industry', 'General', 'Single Post',
     'What happens when you cant pig a pipeline?', 'LOW', 'N'],
    ['Coating Disbondment Detection', 'Educational', 'EMAT', 'Carousel',
     'When coating fails the pipe is exposed. Can you detect it?', 'MEDIUM', 'N'],
    ['Re-inspection Intervals', 'Educational', 'UTWM+UTCD', 'Single Post',
     'How often should you inspect? Depends on growth rate.', 'MEDIUM', 'N'],
]

for row_idx, row_data in enumerate(topics, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws5.cell(row=row_idx, column=col_idx, value=val)
    style_row(ws5, row_idx, len(tb_headers), alt=(row_idx % 2 == 0))

pri_dv = DataValidation(type='list', formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
ws5.add_data_validation(pri_dv)
for r in range(2, 102):
    pri_dv.add(ws5.cell(row=r, column=6))

ws5.freeze_panes = 'A2'

wb.save(r'C:\Outreach\Content-Tracker.xlsx')
print('Content-Tracker.xlsx created successfully with 5 sheets')
